from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify, session, g
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import sqlite3
import os
import logging
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from functools import wraps

app = Flask(__name__)
app.secret_key = '241286'  # Troque por algo único e seguro (ex.: 'xyz123abc')

# Configurar logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Data atual fixa para exemplo (30/03/2025)
DATA_ATUAL = datetime(2025, 3, 30)

# Função para conectar ao banco de dados
def get_db_connection():
    conn = sqlite3.connect('festa.db')
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA busy_timeout = 30000')  # Aumentar o timeout para 30 segundos
    return conn

def get_total_parcelas(numero, transacao_id):
    conn = get_db_connection()
    total_parcelas = conn.execute('SELECT parcelas FROM transacoes WHERE id = ?', (transacao_id,)).fetchone()
    conn.close()
    return total_parcelas['parcelas'] if total_parcelas else 0

# Registrar o filtro no Jinja2
app.jinja_env.filters['get_total_parcelas'] = get_total_parcelas

# Decorador para exigir login
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Decorador para exigir ROOT (opcional, caso queira usar em outras rotas)
def root_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or not session.get('is_root'):
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# Carregar usuário logado antes de cada requisição
@app.before_request
def load_logged_in_user():
    user_id = session.get('user_id')
    if user_id is None:
        g.user = None
    else:
        conn = get_db_connection()
        g.user = conn.execute('SELECT * FROM usuarios WHERE id = ?', (user_id,)).fetchone()
        conn.close()

# Inicializa o banco de dados
def init_db():
    conn = get_db_connection()
    # Tabela de usuários
    conn.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL,
            is_root INTEGER NOT NULL DEFAULT 0
        )
    ''')
    # Tabelas existentes
    conn.execute('''
        CREATE TABLE IF NOT EXISTS participantes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            adulto INTEGER NOT NULL,
            crianca INTEGER NOT NULL,
            bebe_alcool INTEGER NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS transacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT NOT NULL,
            descricao TEXT NOT NULL,
            valor REAL NOT NULL,
            parcelas INTEGER NOT NULL,
            participante_id INTEGER,
            orcamento_id INTEGER,
            FOREIGN KEY (participante_id) REFERENCES participantes(id),
            FOREIGN KEY (orcamento_id) REFERENCES orcamentos(id)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS parcelas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            transacao_id INTEGER,
            numero INTEGER NOT NULL,
            valor REAL NOT NULL,
            data_vencimento TEXT NOT NULL,
            pago INTEGER NOT NULL,
            data_pagamento TEXT,
            FOREIGN KEY (transacao_id) REFERENCES transacoes(id)
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS orcamentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            valor_previsto REAL NOT NULL,
            categoria TEXT
        )
    ''')
    # Verificar e adicionar colunas apenas se necessário
    cursor = conn.execute('PRAGMA table_info(transacoes)')
    transacoes_columns = [col[1] for col in cursor.fetchall()]
    if 'orcamento_id' not in transacoes_columns:
        conn.execute('ALTER TABLE transacoes ADD COLUMN orcamento_id INTEGER')

    cursor = conn.execute('PRAGMA table_info(parcelas)')
    parcelas_columns = [col[1] for col in cursor.fetchall()]
    if 'data_pagamento' not in parcelas_columns:
        conn.execute('ALTER TABLE parcelas ADD COLUMN data_pagamento TEXT')
    
    conn.commit()
    conn.close()

# Rota de login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        with get_db_connection() as conn:
            try:
                user = conn.execute('SELECT * FROM usuarios WHERE username = ? AND password = ?', (username, password)).fetchone()
                if user:
                    session['user_id'] = user['id']
                    session['is_root'] = bool(user['is_root'])
                    return redirect(url_for('index'))
                else:
                    return render_template('login.html', erro='Usuário ou senha inválidos')
            except sqlite3.Error as e:
                logger.error(f"Erro no login: {e}")
                return "Erro no banco de dados!", 500
    return render_template('login.html')

# Rota de cadastro de usuários
@app.route('/cadastro_usuario', methods=['GET', 'POST'])
@login_required
def cadastro_usuario():
    if not session.get('is_root'):  # Apenas ROOT pode cadastrar
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        with get_db_connection() as conn:
            try:
                existing_user = conn.execute('SELECT * FROM usuarios WHERE username = ?', (username,)).fetchone()
                if existing_user:
                    return render_template('cadastro_usuario.html', erro='Usuário já existe', is_root=session.get('is_root', False))
                conn.execute('INSERT INTO usuarios (username, password, is_root) VALUES (?, ?, 0)', (username, password))
                conn.commit()
                return render_template('cadastro_usuario.html', sucesso='Usuário cadastrado com sucesso!', is_root=session.get('is_root', False))
            except sqlite3.Error as e:
                conn.rollback()
                logger.error(f"Erro no cadastro: {e}")
                return "Erro no banco de dados!", 500
    return render_template('cadastro_usuario.html', is_root=session.get('is_root', False))

# ... (código anterior do app.py mantido)

# Rota para alterar senha
@app.route('/alterar_senha', methods=['GET', 'POST'])
@login_required
def alterar_senha():
    if request.method == 'POST':
        senha_atual = request.form['senha_atual']
        nova_senha = request.form['nova_senha']
        confirmar_senha = request.form['confirmar_senha']
        
        with get_db_connection() as conn:
            try:
                user = conn.execute('SELECT * FROM usuarios WHERE id = ?', (session['user_id'],)).fetchone()
                if user['password'] != senha_atual:
                    return render_template('alterar_senha.html', erro='Senha atual incorreta', is_root=session.get('is_root', False))
                if nova_senha != confirmar_senha:
                    return render_template('alterar_senha.html', erro='As novas senhas não coincidem', is_root=session.get('is_root', False))
                if len(nova_senha) < 6:
                    return render_template('alterar_senha.html', erro='A nova senha deve ter pelo menos 6 caracteres', is_root=session.get('is_root', False))
                
                conn.execute('UPDATE usuarios SET password = ? WHERE id = ?', (nova_senha, session['user_id']))
                conn.commit()
                return render_template('alterar_senha.html', sucesso='Senha alterada com sucesso!', is_root=session.get('is_root', False))
            except sqlite3.Error as e:
                conn.rollback()
                logger.error(f"Erro ao alterar senha: {e}")
                return "Erro no banco de dados!", 500
    
    return render_template('alterar_senha.html', is_root=session.get('is_root', False))

# ... (restante do código mantido)

# Rota de logout
@app.route('/logout')
@login_required
def logout():
    session.pop('user_id', None)
    session.pop('is_root', None)
    return redirect(url_for('login'))

# Rota principal (dashboard)
@app.route('/')
@login_required
def index():
    data_atual = datetime.now().strftime('%Y-%m-%d')
    with get_db_connection() as conn:
        try:
            entradas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "entrada"').fetchone()[0] or 0
            saidas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "saida"').fetchone()[0] or 0
            participantes = conn.execute('SELECT COUNT(*) FROM participantes').fetchone()[0]
            orcamentos = conn.execute('SELECT COUNT(*) FROM orcamentos').fetchone()[0]
            transacoes_count = conn.execute('SELECT COUNT(*) FROM transacoes').fetchone()[0]
            parcelas_vencidas = conn.execute('SELECT COUNT(*) FROM parcelas WHERE data_vencimento < ? AND data_pagamento IS NULL', (data_atual,)).fetchone()[0]
            parcelas_a_vencer = conn.execute('SELECT COUNT(*) FROM parcelas WHERE data_vencimento >= ? AND data_pagamento IS NULL', (data_atual,)).fetchone()[0]

            entradas_por_mes = {}
            saidas_por_mes = {}
            transacoes = conn.execute('''
                SELECT t.tipo, t.valor, t.id AS transacao_id, strftime("%Y-%m", p.data_vencimento) AS mes 
                FROM transacoes t 
                JOIN parcelas p ON t.id = p.transacao_id
            ''').fetchall()
            for t in transacoes:
                mes = t['mes']
                valor = t['valor'] / conn.execute('SELECT parcelas FROM transacoes WHERE id = ?', (t['transacao_id'],)).fetchone()['parcelas']
                if t['tipo'] == 'entrada':
                    entradas_por_mes[mes] = entradas_por_mes.get(mes, 0) + valor
                else:
                    saidas_por_mes[mes] = saidas_por_mes.get(mes, 0) + valor

            return render_template('index.html', 
                                   entradas=entradas, 
                                   saidas=saidas, 
                                   participantes=participantes, 
                                   orcamentos=orcamentos, 
                                   transacoes_count=transacoes_count,
                                   parcelas_vencidas=parcelas_vencidas, 
                                   parcelas_a_vencer=parcelas_a_vencer, 
                                   entradas_por_mes=entradas_por_mes, 
                                   saidas_por_mes=saidas_por_mes,
                                   is_root=session.get('is_root', False))  # Certifique-se de que isso está aqui
        except sqlite3.Error as e:
            logger.error(f"Erro no index: {e}")
            return "Erro no banco de dados!", 500

# Controle de participantes
@app.route('/participantes', methods=['GET', 'POST'])
@login_required
def controle_participantes():
    with get_db_connection() as conn:
        try:
            if request.method == 'POST':
                if not session.get('is_root'):  # Apenas ROOT pode criar
                    return "Permissão negada: apenas administradores podem criar participantes.", 403
                nome = request.form['nome']
                adulto = 1 if request.form.get('adulto') == 'on' else 0
                crianca = 1 if request.form.get('crianca') == 'on' else 0
                bebe_alcool = 1 if request.form.get('bebe_alcool') == 'on' else 0

                if adulto and crianca:
                    return "Erro: Uma pessoa não pode ser adulto e criança ao mesmo tempo!", 400
                if crianca and bebe_alcool:
                    return "Erro: Crianças não podem consumir bebida alcoólica!", 400

                conn.execute('INSERT INTO participantes (nome, adulto, crianca, bebe_alcool) VALUES (?, ?, ?, ?)',
                             (nome, adulto, crianca, bebe_alcool))
                conn.commit()
            
            participantes = conn.execute('SELECT * FROM participantes').fetchall()
            return render_template('participantes.html', participantes=participantes, is_root=session.get('is_root', False))
        except sqlite3.Error as e:
            conn.rollback()
            logger.error(f"Erro em participantes: {e}")
            return "Erro no banco de dados!", 500

# Importar participantes
@app.route('/importar_participantes', methods=['POST'])
@login_required
def importar_participantes():
    if not session.get('is_root'):  # Apenas ROOT pode importar
        return "Permissão negada: apenas administradores podem importar participantes.", 403
    
    if 'arquivo_excel' not in request.files:
        return "Nenhum arquivo enviado!", 400
    
    arquivo = request.files['arquivo_excel']
    if arquivo.filename == '':
        return "Nenhum arquivo selecionado!", 400
    
    if not arquivo.filename.endswith('.xlsx'):
        return "Formato inválido! Use um arquivo .xlsx.", 400

    try:
        wb = load_workbook(arquivo)
        ws = wb.active
        expected_headers = ["Nome", "Adulto", "Criança", "Bebe Álcool"]
        headers = [cell.value for cell in ws[1]]
        if headers != expected_headers:
            return "Formato inválido! O arquivo deve conter as colunas: Nome, Adulto, Criança, Bebe Álcool.", 400

        with get_db_connection() as conn:
            for row in ws.iter_rows(min_row=2, values_only=True):
                nome, adulto, crianca, bebe_alcool = row
                if not nome or not isinstance(nome, str):
                    continue
                adulto = 1 if adulto in (1, True, "1", "sim") else 0
                crianca = 1 if crianca in (1, True, "1", "sim") else 0
                bebe_alcool = 1 if bebe_alcool in (1, True, "1", "sim") else 0

                if adulto and crianca:
                    continue
                if crianca and bebe_alcool:
                    continue

                existente = conn.execute('SELECT id FROM participantes WHERE nome = ?', (nome,)).fetchone()
                if not existente:
                    conn.execute('INSERT INTO participantes (nome, adulto, crianca, bebe_alcool) VALUES (?, ?, ?, ?)',
                                 (nome, adulto, crianca, bebe_alcool))
            conn.commit()
        return redirect(url_for('controle_participantes'))
    
    except sqlite3.Error as e:
        logger.error(f"Erro ao importar participantes: {e}")
        return "Erro no banco de dados!", 500
    except Exception as e:
        return f"Erro ao processar o arquivo: {str(e)}", 500

# Editar participante
@app.route('/editar_participante/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_participante(id):
    with get_db_connection() as conn:
        try:
            participante = conn.execute('SELECT * FROM participantes WHERE id = ?', (id,)).fetchone()
            if not participante:
                return "Participante não encontrado!", 404

            if request.method == 'POST':
                if not session.get('is_root'):  # Apenas ROOT pode editar
                    return "Permissão negada: apenas administradores podem editar participantes.", 403
                nome = request.form['nome']
                adulto = 1 if request.form.get('adulto') == 'on' else 0
                crianca = 1 if request.form.get('crianca') == 'on' else 0
                bebe_alcool = 1 if request.form.get('bebe_alcool') == 'on' else 0

                if adulto and crianca:
                    return "Erro: Uma pessoa não pode ser adulto e criança ao mesmo tempo!", 400
                if crianca and bebe_alcool:
                    return "Erro: Crianças não podem consumir bebida alcoólica!", 400

                conn.execute('UPDATE participantes SET nome = ?, adulto = ?, crianca = ?, bebe_alcool = ? WHERE id = ?',
                             (nome, adulto, crianca, bebe_alcool, id))
                conn.commit()
                return redirect(url_for('controle_participantes'))

            participantes = conn.execute('SELECT * FROM participantes').fetchall()
            return render_template('participantes.html', participantes=participantes, edit_participante=participante, is_root=session.get('is_root', False))
        except sqlite3.Error as e:
            conn.rollback()
            logger.error(f"Erro em editar_participante: {e}")
            return "Erro no banco de dados!", 500

# Excluir participante
@app.route('/excluir_participante/<int:id>')
@login_required
def excluir_participante(id):
    if not session.get('is_root'):  # Apenas ROOT pode excluir
        return "Permissão negada: apenas administradores podem excluir participantes.", 403
    
    with get_db_connection() as conn:
        try:
            conn.execute('DELETE FROM participantes WHERE id = ?', (id,))
            conn.commit()
            return redirect(url_for('controle_participantes'))
        except sqlite3.Error as e:
            conn.rollback()
            logger.error(f"Erro em excluir_participante: {e}")
            return "Erro no banco de dados!", 500
        
        
# Controle financeiro
@app.route('/financeiro', methods=['GET', 'POST'])
@login_required
def controle_financeiro():
    data_atual = datetime.now().strftime('%Y-%m-%d')
    with get_db_connection() as conn:
        try:
            if request.method == 'POST':
                if not session.get('is_root'):  # Apenas ROOT pode modificar
                    return "Permissão negada: apenas administradores podem modificar transações.", 403
                action = request.form.get('action')
                if action == 'cadastrar':
                    tipo = request.form['tipo']
                    descricao = request.form['descricao']
                    participante_id = request.form['participante_id']
                    valor = float(request.form['valor'])
                    parcelas = int(request.form['parcelas'])
                    data_vencimento = request.form['data_vencimento']
                    orcamento_id = request.form.get('orcamento_id')  # Novo campo

                    conn.execute('INSERT INTO transacoes (tipo, descricao, participante_id, valor, parcelas, orcamento_id) VALUES (?, ?, ?, ?, ?, ?)',
                                 (tipo, descricao, participante_id, valor, parcelas, orcamento_id or None))
                    transacao_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
                    valor_parcela = valor / parcelas
                    data_inicial = datetime.strptime(data_vencimento, '%Y-%m-%d')
                    for i in range(parcelas):
                        data_parcela = (data_inicial + timedelta(days=30 * i)).strftime('%Y-%m-%d')
                        conn.execute('INSERT INTO parcelas (transacao_id, numero, valor, data_vencimento, pago) VALUES (?, ?, ?, ?, ?)',
                                     (transacao_id, i + 1, valor_parcela, data_parcela, 0))
                    conn.commit()

                    participante_nome = conn.execute('SELECT nome FROM participantes WHERE id = ?', (participante_id,)).fetchone()['nome']
                    return jsonify({
                        'success': True,
                        'transacao_id': transacao_id,
                        'tipo': tipo,
                        'descricao': descricao,
                        'participante_id': participante_id,
                        'participante_nome': participante_nome,
                        'valor': valor,
                        'parcelas': parcelas,
                        'data_vencimento': data_vencimento,
                        'orcamento_id': orcamento_id
                    })
                elif action == 'editar':
                    transacao_id = request.form['transacao_id']
                    tipo = request.form['tipo']
                    descricao = request.form['descricao']
                    participante_id = request.form['participante_id']
                    valor = float(request.form['valor'])
                    parcelas = int(request.form['parcelas'])
                    data_vencimento = request.form['data_vencimento']
                    orcamento_id = request.form.get('orcamento_id')  # Novo campo

                    conn.execute('UPDATE transacoes SET tipo = ?, descricao = ?, participante_id = ?, valor = ?, parcelas = ?, orcamento_id = ? WHERE id = ?',
                                 (tipo, descricao, participante_id, valor, parcelas, orcamento_id or None, transacao_id))
                    conn.execute('DELETE FROM parcelas WHERE transacao_id = ?', (transacao_id,))
                    valor_parcela = valor / parcelas
                    data_inicial = datetime.strptime(data_vencimento, '%Y-%m-%d')
                    for i in range(parcelas):
                        data_parcela = (data_inicial + timedelta(days=30 * i)).strftime('%Y-%m-%d')
                        conn.execute('INSERT INTO parcelas (transacao_id, numero, valor, data_vencimento, pago) VALUES (?, ?, ?, ?, ?)',
                                     (transacao_id, i + 1, valor_parcela, data_parcela, 0))
                    conn.commit()
                elif action == 'excluir':
                    transacao_id = request.form['transacao_id']
                    conn.execute('DELETE FROM parcelas WHERE transacao_id = ?', (transacao_id,))
                    conn.execute('DELETE FROM transacoes WHERE id = ?', (transacao_id,))
                    conn.commit()
                return redirect(url_for('controle_financeiro', busca=request.args.get('busca', '')))

            busca = request.args.get('busca', '')
            query = '''
                SELECT t.id, t.tipo, t.descricao, t.valor, t.parcelas, par.data_vencimento, p.nome AS participante_nome, t.participante_id, t.orcamento_id, o.nome AS orcamento_nome
                FROM transacoes t
                JOIN participantes p ON t.participante_id = p.id
                JOIN parcelas par ON t.id = par.transacao_id
                LEFT JOIN orcamentos o ON t.orcamento_id = o.id
                WHERE par.numero = 1
            '''
            params = []
            if busca:
                query += ' AND t.descricao LIKE ?'
                params.append(f'%{busca}%')
            fluxo_caixa = conn.execute(query, params).fetchall()
            participantes = conn.execute('SELECT id, nome FROM participantes').fetchall()
            orcamentos = conn.execute('SELECT id, nome FROM orcamentos').fetchall()

            return render_template('financeiro.html', fluxo_caixa=fluxo_caixa, busca=busca, participantes=participantes, 
                                   orcamentos=orcamentos, data_atual=data_atual, is_root=session.get('is_root', False))
        except sqlite3.Error as e:
            conn.rollback()
            logger.error(f"Erro em financeiro: {e}")
            return "Erro no banco de dados!", 500

# Editar transação
@app.route('/editar_transacao/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_transacao(id):
    with get_db_connection() as conn:
        try:
            transacao = conn.execute('SELECT t.*, p.nome AS participante_nome FROM transacoes t JOIN participantes p ON t.participante_id = p.id WHERE t.id = ?', (id,)).fetchone()
            if not transacao:
                return "Transação não encontrada!", 404

            if request.method == 'POST':
                if not session.get('is_root'):  # Apenas ROOT pode editar
                    return "Permissão negada: apenas administradores podem editar transações.", 403
                novo_numero_parcelas = int(request.form['parcelas'])
                valor_total = float(transacao['valor'])
                data_vencimento_inicial = datetime.fromisoformat(conn.execute('SELECT data_vencimento FROM parcelas WHERE transacao_id = ? AND numero = 1', (id,)).fetchone()['data_vencimento'])
                
                novo_valor_parcela = round(valor_total / novo_numero_parcelas, 2)
                conn.execute('DELETE FROM parcelas WHERE transacao_id = ?', (id,))
                for i in range(novo_numero_parcelas):
                    vencimento = data_vencimento_inicial + timedelta(days=30 * i)
                    conn.execute('INSERT INTO parcelas (transacao_id, numero, valor, data_vencimento, pago) VALUES (?, ?, ?, ?, ?)',
                                 (id, i + 1, novo_valor_parcela, vencimento.isoformat(), 0))
                conn.execute('UPDATE transacoes SET parcelas = ? WHERE id = ?', (novo_numero_parcelas, id))
                
                conn.commit()
                return redirect(url_for('controle_financeiro'))

            participantes = conn.execute('SELECT * FROM participantes').fetchall()
            fluxo = conn.execute('SELECT t.*, p.nome AS participante_nome FROM transacoes t JOIN participantes p ON t.participante_id = p.id').fetchall()
            return render_template('financeiro.html', fluxo=fluxo, participantes=participantes, resumo_parcelas={}, edit_transacao=transacao, is_root=session.get('is_root', False))
        except sqlite3.Error as e:
            conn.rollback()
            logger.error(f"Erro em editar_transacao: {e}")
            return "Erro no banco de dados!", 500

# Dar baixa em parcelas
@app.route('/dar_baixa_parcelas', methods=['POST'])
@login_required
def dar_baixa_parcelas():
    if not session.get('is_root'):  # Apenas ROOT pode dar baixa
        return "Permissão negada: apenas administradores podem dar baixa em parcelas.", 403
    
    with get_db_connection() as conn:
        try:
            parcelas_selecionadas = request.form.getlist('parcelas[]')
            
            for parcela_id in parcelas_selecionadas:
                id_transacao, numero = map(int, parcela_id.split('-'))
                data_pagamento = request.form.get(f'data_pagamento_{id_transacao}-{numero}')
                if data_pagamento:
                    conn.execute('''
                        UPDATE parcelas 
                        SET pago = 1, data_pagamento = ? 
                        WHERE transacao_id = ? AND numero = ?
                    ''', (data_pagamento, id_transacao, numero))
            
            conn.commit()
            return redirect(url_for('controle_financeiro'))
        except sqlite3.Error as e:
            conn.rollback()
            logger.error(f"Erro em dar_baixa_parcelas: {e}")
            return "Erro no banco de dados!", 500
        
        
# Editar parcela
@app.route('/editar_parcela/<int:id_transacao>/<int:numero>', methods=['GET', 'POST'])
@login_required
def editar_parcela(id_transacao, numero):
    with get_db_connection() as conn:
        try:
            parcela = conn.execute('SELECT * FROM parcelas WHERE transacao_id = ? AND numero = ?', (id_transacao, numero)).fetchone()
            if not parcela:
                return "Parcela não encontrada!", 404

            transacao = conn.execute('SELECT * FROM transacoes WHERE id = ?', (id_transacao,)).fetchone()
            if not transacao:
                return "Transação não encontrada!", 404

            if request.method == 'POST':
                if not session.get('is_root'):  # Apenas ROOT pode editar
                    return "Permissão negada: apenas administradores podem editar parcelas.", 403
                valor = round(float(request.form['valor']), 2)
                data_vencimento = datetime.strptime(request.form['data_vencimento'], '%Y-%m-%d')
                pago = 1 if request.form.get('pago') == 'on' else 0
                data_pagamento = request.form['data_pagamento'] if pago and request.form['data_pagamento'] else None

                conn.execute('UPDATE parcelas SET valor = ?, data_vencimento = ?, pago = ?, data_pagamento = ? WHERE transacao_id = ? AND numero = ?',
                             (valor, data_vencimento.isoformat(), pago, data_pagamento, id_transacao, numero))
                conn.commit()
                return redirect(url_for('controle_financeiro'))

            participantes = conn.execute('SELECT * FROM participantes').fetchall()
            fluxo = conn.execute('SELECT t.*, p.nome AS participante_nome FROM transacoes t JOIN participantes p ON t.participante_id = p.id').fetchall()
            return render_template('financeiro.html', fluxo=fluxo, participantes=participantes, resumo_parcelas={}, 
                                   edit_parcela={'transacao': transacao, 'parcela': parcela}, is_root=session.get('is_root', False))
        except sqlite3.Error as e:
            conn.rollback()
            logger.error(f"Erro em editar_parcela: {e}")
            return "Erro no banco de dados!", 500

# Excluir parcela
@app.route('/excluir_parcela/<int:id_transacao>/<int:numero>')
@login_required
def excluir_parcela(id_transacao, numero):
    if not session.get('is_root'):  # Apenas ROOT pode excluir
        return "Permissão negada: apenas administradores podem excluir parcelas.", 403
    
    with get_db_connection() as conn:
        try:
            conn.execute('DELETE FROM parcelas WHERE transacao_id = ? AND numero = ?', (id_transacao, numero))
            
            parcelas_restantes = conn.execute('SELECT COUNT(*) FROM parcelas WHERE transacao_id = ?', (id_transacao,)).fetchone()[0]
            conn.execute('UPDATE transacoes SET parcelas = ? WHERE id = ?', (parcelas_restantes, id_transacao))
            
            if parcelas_restantes == 0:
                conn.execute('DELETE FROM transacoes WHERE id = ?', (id_transacao,))
            
            conn.commit()
            return redirect(url_for('controle_financeiro'))
        except sqlite3.Error as e:
            conn.rollback()
            logger.error(f"Erro em excluir_parcela: {e}")
            return "Erro no banco de dados!", 500

# Resumo financeiro
@app.route('/resumo')
@login_required
def resumo_financeiro():
    conn = get_db_connection()
    total_entradas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "entrada"').fetchone()[0] or 0
    total_saidas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "saida"').fetchone()[0] or 0
    saldo = total_entradas - total_saidas

    parcelas = conn.execute('SELECT valor, data_vencimento, pago FROM parcelas').fetchall()
    valores_recebidos = 0
    valores_a_vencer = 0
    valores_vencidos = 0
    
    for parcela in parcelas:
        data_vencimento = datetime.fromisoformat(parcela['data_vencimento'])
        if parcela['pago']:
            valores_recebidos += parcela['valor']
        elif data_vencimento < DATA_ATUAL:
            valores_vencidos += parcela['valor']
        else:
            valores_a_vencer += parcela['valor']

    participantes = conn.execute('SELECT * FROM participantes').fetchall()
    resumo_participantes = {}
    
    for p in participantes:
        entradas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "entrada" AND participante_id = ?', (p['id'],)).fetchone()[0] or 0
        saidas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "saida" AND participante_id = ?', (p['id'],)).fetchone()[0] or 0
        
        parcelas = conn.execute('''
            SELECT par.valor, par.data_vencimento, par.pago
            FROM transacoes t
            JOIN parcelas par ON t.id = par.transacao_id
            WHERE t.participante_id = ?
        ''', (p['id'],)).fetchall()
        
        p_recebidos = 0
        p_a_vencer = 0
        p_vencidos = 0
        
        for parcela in parcelas:
            data_vencimento = datetime.fromisoformat(parcela['data_vencimento'])
            if parcela['pago']:
                p_recebidos += parcela['valor']
            elif data_vencimento < DATA_ATUAL:
                p_vencidos += parcela['valor']
            else:
                p_a_vencer += parcela['valor']
        
        resumo_participantes[p['nome']] = {
            'entradas': entradas,
            'saidas': saidas,
            'valores_recebidos': p_recebidos,
            'valores_a_vencer': p_a_vencer,
            'valores_vencidos': p_vencidos
        }

    resumo = {
        'total_entradas': total_entradas,
        'total_saidas': total_saidas,
        'saldo': saldo,
        'valores_recebidos': valores_recebidos,
        'valores_a_vencer': valores_a_vencer,
        'valores_vencidos': valores_vencidos,
        'participantes': resumo_participantes
    }
    conn.close()
    return render_template('resumo.html', resumo=resumo, is_root=session.get('is_root', False))

# Gerar PDF
@app.route('/gerar_pdf')
@login_required
def gerar_pdf():
    conn = get_db_connection()
    total_entradas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "entrada"').fetchone()[0] or 0
    total_saidas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "saida"').fetchone()[0] or 0
    saldo = total_entradas - total_saidas

    parcelas = conn.execute('SELECT valor, data_vencimento, pago FROM parcelas').fetchall()
    valores_recebidos = 0
    valores_a_vencer = 0
    valores_vencidos = 0
    
    for parcela in parcelas:
        data_vencimento = datetime.fromisoformat(parcela['data_vencimento'])
        if parcela['pago']:
            valores_recebidos += parcela['valor']
        elif data_vencimento < DATA_ATUAL:
            valores_vencidos += parcela['valor']
        else:
            valores_a_vencer += parcela['valor']

    participantes = conn.execute('SELECT * FROM participantes').fetchall()
    resumo_participantes = {}
    
    for p in participantes:
        entradas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "entrada" AND participante_id = ?', (p['id'],)).fetchone()[0] or 0
        saidas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "saida" AND participante_id = ?', (p['id'],)).fetchone()[0] or 0
        
        parcelas = conn.execute('''
            SELECT par.valor, par.data_vencimento, par.pago
            FROM transacoes t
            JOIN parcelas par ON t.id = par.transacao_id
            WHERE t.participante_id = ?
        ''', (p['id'],)).fetchall()
        
        p_recebidos = 0
        p_a_vencer = 0
        p_vencidos = 0
        
        for parcela in parcelas:
            data_vencimento = datetime.fromisoformat(parcela['data_vencimento'])
            if parcela['pago']:
                p_recebidos += parcela['valor']
            elif data_vencimento < DATA_ATUAL:
                p_vencidos += parcela['valor']
            else:
                p_a_vencer += parcela['valor']
        
        resumo_participantes[p['nome']] = {
            'entradas': entradas,
            'saidas': saidas,
            'valores_recebidos': p_recebidos,
            'valores_a_vencer': p_a_vencer,
            'valores_vencidos': p_vencidos
        }

    conn.close()

    pdf_file = "resumo_financeiro.pdf"
    doc = SimpleDocTemplate(pdf_file, pagesize=letter)
    elements = []

    base_styles = getSampleStyleSheet()
    styles = {
        'Title': ParagraphStyle('Title', parent=base_styles['Title'], fontName='Helvetica-Bold', fontSize=20, textColor=colors.HexColor('#2c3e50')),
        'Heading2': ParagraphStyle('Heading2', parent=base_styles['Normal'], fontName='Helvetica-Bold', fontSize=14, textColor=colors.HexColor('#2c3e50')),
        'Normal': ParagraphStyle('Normal', parent=base_styles['Normal'], fontName='Helvetica', fontSize=10)
    }

    title = Paragraph("Resumo Financeiro", styles['Title'])
    elements.append(title)
    elements.append(Paragraph("<br/><br/>", styles['Normal']))

    data_totais = [
        ["Entradas", "Saídas", "Saldo", "Recebidos", "A Vencer", "Vencidos"],
        [f"R$ {total_entradas:.2f}", f"R$ {total_saidas:.2f}", f"R$ {saldo:.2f}", 
         f"R$ {valores_recebidos:.2f}", f"R$ {valores_a_vencer:.2f}", f"R$ {valores_vencidos:.2f}"]
    ]
    tabela_totais = Table(data_totais, colWidths=[100, 100, 100, 100, 100, 100])
    tabela_totais.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#27ae60')),
        ('BACKGROUND', (1, 1), (1, 1), colors.HexColor('#e74c3c')),
        ('BACKGROUND', (2, 1), (2, 1), colors.HexColor('#3498db')),
        ('BACKGROUND', (3, 1), (3, 1), colors.HexColor('#27ae60')),
        ('BACKGROUND', (4, 1), (4, 1), colors.HexColor('#e67e22')),
        ('BACKGROUND', (5, 1), (5, 1), colors.HexColor('#e74c3c')),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 2, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(tabela_totais)
    elements.append(Paragraph("<br/><br/>", styles['Normal']))

    elements.append(Paragraph("Resumo por Participante", styles['Heading2']))
    elements.append(Paragraph("<br/>", styles['Normal']))

    data_participantes = [["Participante", "Entradas", "Saídas", "Recebidos", "A Vencer", "Vencidos"]]
    for nome, dados in resumo_participantes.items():
        data_participantes.append([
            nome,
            f"R$ {dados['entradas']:.2f}",
            f"R$ {dados['saidas']:.2f}",
            f"R$ {dados['valores_recebidos']:.2f}",
            f"R$ {dados['valores_a_vencer']:.2f}",
            f"R$ {dados['valores_vencidos']:.2f}"
        ])
    
    tabela_participantes = Table(data_participantes, colWidths=[150, 90, 90, 90, 90, 90])
    tabela_participantes.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fbfd')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 2, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(tabela_participantes)

    doc.build(elements)
    return send_file(pdf_file, as_attachment=True)

# Gerar Excel
@app.route('/gerar_excel')
@login_required
def gerar_excel():
    conn = get_db_connection()
    total_entradas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "entrada"').fetchone()[0] or 0
    total_saidas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "saida"').fetchone()[0] or 0
    saldo = total_entradas - total_saidas

    parcelas = conn.execute('SELECT valor, data_vencimento, pago FROM parcelas').fetchall()
    valores_recebidos = 0
    valores_a_vencer = 0
    valores_vencidos = 0
    
    for parcela in parcelas:
        data_vencimento = datetime.fromisoformat(parcela['data_vencimento'])
        if parcela['pago']:
            valores_recebidos += parcela['valor']
        elif data_vencimento < DATA_ATUAL:
            valores_vencidos += parcela['valor']
        else:
            valores_a_vencer += parcela['valor']

    participantes = conn.execute('SELECT * FROM participantes').fetchall()
    resumo_participantes = {}
    
    for p in participantes:
        entradas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "entrada" AND participante_id = ?', (p['id'],)).fetchone()[0] or 0
        saidas = conn.execute('SELECT SUM(valor) FROM transacoes WHERE tipo = "saida" AND participante_id = ?', (p['id'],)).fetchone()[0] or 0
        
        parcelas = conn.execute('''
            SELECT par.valor, par.data_vencimento, par.pago
            FROM transacoes t
            JOIN parcelas par ON t.id = par.transacao_id
            WHERE t.participante_id = ?
        ''', (p['id'],)).fetchall()
        
        p_recebidos = 0
        p_a_vencer = 0
        p_vencidos = 0
        
        for parcela in parcelas:
            data_vencimento = datetime.fromisoformat(parcela['data_vencimento'])
            if parcela['pago']:
                p_recebidos += parcela['valor']
            elif data_vencimento < DATA_ATUAL:
                p_vencidos += parcela['valor']
            else:
                p_a_vencer += parcela['valor']
        
        resumo_participantes[p['nome']] = {
            'entradas': entradas,
            'saidas': saidas,
            'valores_recebidos': p_recebidos,
            'valores_a_vencer': p_a_vencer,
            'valores_vencidos': p_vencidos
        }

    conn.close()

    wb = Workbook()
    ws_geral = wb.active
    ws_geral.title = "Resumo Geral"
    headers = ["Entradas", "Saídas", "Saldo", "Recebidos", "A Vencer", "Vencidos"]
    values = [total_entradas, total_saidas, saldo, valores_recebidos, valores_a_vencer, valores_vencidos]

    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws_geral.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    value_fill = PatternFill(start_color="f9fbfd", end_color="f9fbfd", fill_type="solid")
    for col, value in enumerate(values, 1):
        cell = ws_geral.cell(row=2, column=col, value=f"R$ {value:.2f}")
        cell.fill = value_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, 7):
        ws_geral.column_dimensions[ws_geral.cell(row=1, column=col).column_letter].width = 15

    ws_participantes = wb.create_sheet("Por Participante")
    headers = ["Participante", "Entradas", "Saídas", "Recebidos", "A Vencer", "Vencidos"]
    for col, header in enumerate(headers, 1):
        cell = ws_participantes.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, (nome, dados) in enumerate(resumo_participantes.items(), 2):
        ws_participantes.cell(row=row, column=1, value=nome).fill = value_fill
        ws_participantes.cell(row=row, column=2, value=f"R$ {dados['entradas']:.2f}").fill = value_fill
        ws_participantes.cell(row=row, column=3, value=f"R$ {dados['saidas']:.2f}").fill = value_fill
        ws_participantes.cell(row=row, column=4, value=f"R$ {dados['valores_recebidos']:.2f}").fill = value_fill
        ws_participantes.cell(row=row, column=5, value=f"R$ {dados['valores_a_vencer']:.2f}").fill = value_fill
        ws_participantes.cell(row=row, column=6, value=f"R$ {dados['valores_vencidos']:.2f}").fill = value_fill
        for col in range(1, 7):
            ws_participantes.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    ws_participantes.column_dimensions['A'].width = 20
    for col in range(2, 7):
        ws_participantes.column_dimensions[ws_participantes.cell(row=1, column=col).column_letter].width = 15

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name="resumo_financeiro.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def calcular_resumo_orcamentos(orcamentos):
    conn = get_db_connection()
    resumo = []
    try:
        for orcamento in orcamentos:
            valor_real = conn.execute(
                'SELECT SUM(valor) FROM transacoes WHERE orcamento_id = ? AND tipo = "saida"',
                (orcamento['id'],)
            ).fetchone()[0] or 0
            status = 'Dentro do Orçamento' if valor_real <= orcamento['valor_previsto'] else 'Fora do Orçamento'
            resumo.append({
                'id': orcamento['id'],
                'nome': orcamento['nome'],
                'valor_previsto': orcamento['valor_previsto'],
                'valor_real': valor_real,
                'categoria': orcamento['categoria'],
                'status': status
            })
        conn.close()
        return resumo
    except sqlite3.Error as e:
        conn.rollback()
        logger.error(f"Erro em calcular_resumo_orcamentos: {e}")
        conn.close()
        return []



# Controle de orçamentos
from flask import request, redirect, url_for, render_template, flash, session

@app.route('/orcamentos', methods=['GET', 'POST'])
@login_required
def controle_orcamentos():
    conn = get_db_connection()
    try:
        if request.method == 'POST':
            nome = request.form['nome']
            valor_previsto = float(request.form['valor_previsto'])
            categoria = request.form.get('categoria', '')
            conn.execute('INSERT INTO orcamentos (nome, valor_previsto, categoria) VALUES (?, ?, ?)',
                         (nome, valor_previsto, categoria))
            conn.commit()
            flash('Orçamento cadastrado com sucesso!', 'success')
            return redirect(url_for('controle_orcamentos'))

        orcamentos = conn.execute('SELECT * FROM orcamentos').fetchall()
        resumo_orcamentos = calcular_resumo_orcamentos(orcamentos)
        conn.close()
        return render_template('orcamentos.html', resumo_orcamentos=resumo_orcamentos, is_root=session.get('is_root', False))
    except sqlite3.Error as e:
        conn.rollback()
        logger.error(f"Erro em controle_orcamentos: {e}")
        conn.close()
        return "Erro no banco de dados!", 500

# editar_orçamentos
from flask import request, redirect, url_for, render_template, flash, session

@app.route('/editar_orcamento/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_orcamento(id):
    conn = get_db_connection()
    try:
        orcamento = conn.execute('SELECT * FROM orcamentos WHERE id = ?', (id,)).fetchone()
        if not orcamento:
            conn.close()
            return "Orçamento não encontrado!", 404

        if request.method == 'POST':
            nome = request.form['nome']
            valor_previsto = float(request.form['valor_previsto'])
            categoria = request.form.get('categoria', '')
            conn.execute('UPDATE orcamentos SET nome = ?, valor_previsto = ?, categoria = ? WHERE id = ?',
                         (nome, valor_previsto, categoria, id))
            conn.commit()
            flash('Orçamento editado com sucesso!', 'success')
            conn.close()
            return redirect(url_for('controle_orcamentos'))

        orcamentos = conn.execute('SELECT * FROM orcamentos').fetchall()
        resumo_orcamentos = calcular_resumo_orcamentos(orcamentos)
        conn.close()
        return render_template('orcamentos.html', edit_orcamento=orcamento, resumo_orcamentos=resumo_orcamentos, is_root=session.get('is_root', False))
    except sqlite3.Error as e:
        conn.rollback()
        logger.error(f"Erro em editar_orcamento: {e}")
        conn.close()
        return "Erro no banco de dados!", 500

from flask import request, redirect, url_for, render_template, flash, session

@app.route('/excluir_orcamento/<int:id>')
@login_required
def excluir_orcamento(id):
    if not session.get('is_root'):
        flash('Permissão negada: apenas administradores podem excluir orçamentos.', 'error')
        return redirect(url_for('controle_orcamentos'))
    
    conn = get_db_connection()
    try:
        orcamento = conn.execute('SELECT * FROM orcamentos WHERE id = ?', (id,)).fetchone()
        if not orcamento:
            conn.close()
            flash('Orçamento não encontrado!', 'error')
            return redirect(url_for('controle_orcamentos'))
        
        conn.execute('DELETE FROM orcamentos WHERE id = ?', (id,))
        conn.commit()
        flash('Orçamento excluído com sucesso!', 'success')
        conn.close()
        return redirect(url_for('controle_orcamentos'))
    except sqlite3.Error as e:
        conn.rollback()
        logger.error(f"Erro em excluir_orcamento: {e}")
        conn.close()
        flash('Erro ao excluir o orçamento devido a um problema no banco de dados!', 'error')
        return redirect(url_for('controle_orcamentos'))


# Gerenciar baixas
@app.route('/baixas', methods=['GET', 'POST'])
@login_required
def baixas():
    conn = get_db_connection()
    data_atual = datetime.now().strftime('%Y-%m-%d')

    try:
        if request.method == 'POST':
            if not session.get('is_root'):  # Apenas ROOT pode modificar
                return "Permissão negada: apenas administradores podem gerenciar baixas.", 403
            action = request.form.get('action')
            if action == 'dar_baixa':
                parcela_id = request.form['parcela_id']
                data_pagamento = request.form['data_pagamento']
                conn.execute('UPDATE parcelas SET data_pagamento = ? WHERE id = ?', (data_pagamento, parcela_id))
                conn.commit()
                parcela = conn.execute('SELECT t.id AS transacao_id, t.parcelas AS total_parcelas FROM parcelas p JOIN transacoes t ON p.transacao_id = t.id WHERE p.id = ?', (parcela_id,)).fetchone()
                return jsonify({
                    'success': True,
                    'parcela_id': parcela_id,
                    'data_pagamento': data_pagamento,
                    'transacao_id': parcela['transacao_id'],
                    'total_parcelas': parcela['total_parcelas']
                })
            elif action == 'editar_parcelas':
                transacao_id = request.form['transacao_id']
                novas_parcelas = int(request.form['novas_parcelas'])
                conn.execute('UPDATE transacoes SET parcelas = ? WHERE id = ?', (novas_parcelas, transacao_id))
                conn.commit()
            elif action == 'editar_parcela':
                parcela_id = request.form['parcela_id']
                valor = float(request.form['valor'])
                data_vencimento = request.form['data_vencimento']
                conn.execute('UPDATE parcelas SET valor = ?, data_vencimento = ? WHERE id = ?', (valor, data_vencimento, parcela_id))
                conn.commit()
            elif action == 'editar_baixa':
                parcela_id = request.form['parcela_id']
                data_pagamento = request.form['data_pagamento']
                conn.execute('UPDATE parcelas SET data_pagamento = ? WHERE id = ?', (data_pagamento, parcela_id))
                conn.commit()
            elif action == 'excluir_parcela':
                parcela_id = request.form['parcela_id']
                conn.execute('DELETE FROM parcelas WHERE id = ?', (parcela_id,))
                conn.commit()
            elif action == 'excluir_transacao':
                transacao_id = request.form['transacao_id']
                conn.execute('DELETE FROM parcelas WHERE transacao_id = ?', (transacao_id,))
                conn.execute('DELETE FROM transacoes WHERE id = ?', (transacao_id,))
                conn.commit()
            elif action == 'excluir_baixa':
                parcela_id = request.form['parcela_id']
                conn.execute('UPDATE parcelas SET data_pagamento = NULL WHERE id = ?', (parcela_id,))
                conn.commit()
            conn.close()
            return redirect(url_for('baixas', busca=request.args.get('busca', ''), participante=request.args.get('participante', ''), 
                                   status=request.args.get('status', ''), data_inicio=request.args.get('data_inicio', '')))

        busca = request.args.get('busca', '')
        participante = request.args.get('participante', '')
        status = request.args.get('status', '')
        data_inicio = request.args.get('data_inicio', '')

        query = '''
            SELECT p.id AS parcela_id, t.id AS transacao_id, t.descricao, t.parcelas, p.numero, p.valor, 
                   p.data_vencimento, p.data_pagamento, part.nome AS participante_nome,
                   CASE 
                       WHEN p.data_pagamento IS NOT NULL THEN 'recebidas'
                       WHEN p.data_vencimento < ? THEN 'vencidas'
                       ELSE 'a vencer'
                   END AS status
            FROM parcelas p
            JOIN transacoes t ON p.transacao_id = t.id
            JOIN participantes part ON t.participante_id = part.id
            WHERE 1=1
        '''
        params = [data_atual]

        if busca:
            query += ' AND (t.descricao LIKE ? OR part.nome LIKE ?)'
            params.extend([f'%{busca}%', f'%{busca}%'])
        if participante:
            query += ' AND part.id = ?'
            params.append(participante)
        if status:
            query += ' AND CASE WHEN p.data_pagamento IS NOT NULL THEN "recebidas" WHEN p.data_vencimento < ? THEN "vencidas" ELSE "a vencer" END = ?'
            params.extend([data_atual, status])
        if data_inicio:
            query += ' AND p.data_vencimento >= ?'
            params.append(data_inicio)

        todas_parcelas = conn.execute(query, params).fetchall()
        parcelas_a_vencer = [p for p in todas_parcelas if p['status'] == 'a vencer']
        parcelas_vencidas = [p for p in todas_parcelas if p['status'] == 'vencidas']
        parcelas_recebidas = [p for p in todas_parcelas if p['status'] == 'recebidas']

        participantes = conn.execute('SELECT id, nome FROM participantes').fetchall()
        conn.close()

        return render_template('baixas.html', 
                               parcelas_a_vencer=parcelas_a_vencer, 
                               parcelas_vencidas=parcelas_vencidas, 
                               parcelas_recebidas=parcelas_recebidas, 
                               busca=busca, 
                               participante=participante, 
                               status=status, 
                               data_inicio=data_inicio, 
                               data_atual=data_atual, 
                               participantes=participantes,
                               is_root=session.get('is_root', False))
    except sqlite3.Error as e:
        conn.rollback()
        logger.error(f"Erro em baixas: {e}")
        return "Erro no banco de dados!", 500

# Rota home (substituída pelo index)
@app.route("/home")
def home():
    return "Meu app está online!"

if __name__ == '__main__':
    init_db()
    print("Endpoints registrados:", [rule.endpoint for rule in app.url_map.iter_rules()])
    app.run(debug=True)